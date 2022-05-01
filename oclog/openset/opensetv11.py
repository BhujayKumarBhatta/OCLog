# -*- coding: utf-8 -*-
"""
Created on Sun Feb 13 21:24:31 2022

@author: Bhujay_ROG
"""
import os
import datetime
import time
import uuid
import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np
import pandas as pd
import tensorflow as tf
import seaborn as sns
import matplotlib.pyplot as plt
tf.random.set_seed(123)
from oclog.BGL.bglv1 import BGLog
from oclog.openset.boundary_loss import BoundaryLoss
from oclog.openset.pretrainingV1 import LogLineEncoder, LogSeqEncoder, LogClassifier
from tqdm import trange, tqdm, tnrange
# from time import sleep
from sklearn.metrics import confusion_matrix, f1_score, accuracy_score
import sklearn.metrics as m
import warnings
warnings.filterwarnings('ignore')
from collections import defaultdict
from sklearn.preprocessing import MinMaxScaler, StandardScaler
from sklearn.manifold import TSNE
from tensorflow.keras.callbacks import ModelCheckpoint, ReduceLROnPlateau, EarlyStopping
from tensorflow.keras.models import  load_model



class OpenSet:
    '''  
    init attributes cleaned up
    TODO: update tracker with pttime and octime and notebook name
    TODO: excel cols to be reorganized
    '''
    def __init__(self, function_model=False, ):
        self.ptmodel = None
        self.centroids = None       
        self.radius = None      
        self.radius_changes = []
        self.losses = []
        self.f1_tr_lst = []
        self.f1_val_lst = []
        self.function_model = function_model
        self.batch_features = None
        self.total_features = []
        self.total_preds = []
        self.total_labels = []
        self.epoch = 0
        self.best_train_score = 0
        self.best_val_score = 0        
        self.ptmodel_name = 'ptmodel'
        self.data_dir = 'data'
        self.ptmodel_path = None
        self.num_classes = None       
        self.tf_random_seed = 1234  
        self.tracker = {}
        self.ukc_label = 9
    
    
    def extract_features_and_centroids(self, **kwargs):
        store_features = kwargs.get('store_features', True)
        train_data, val_data,  test_data, bglog = self.get_or_generate_dataset(**kwargs)
        self.centroids = self.centroids_cal(train_data, **kwargs)        
        feature_from = kwargs.get('store_features', 'train_data')        
        if feature_from == 'val_data':
            data = val_data
        elif feature_from == 'test_data':
            data = test_data
        else:
            data = train_data
        total_features, total_preds, total_labels = [], [], []
        for batch in data:
            logseq_batch, label_batch = batch
            features_batch = self.get_pretrained_features(logseq_batch, **kwargs)
            label_indexs = tf.math.argmax(label_batch, axis=1)
            label_index_np = label_indexs.numpy()
            total_labels.append(label_index_np)
            total_features.append(features_batch)
        # y_pred = np.array(total_preds).flatten().tolist()
        y_true = np.array(total_labels).flatten().tolist()
        if store_features:
            # self.total_preds = y_pred
            self.total_labels = y_true
            total_features = np.array(total_features)
            total_features = np.reshape(total_features, ((total_features.shape[0] * total_features.shape[1]), total_features.shape[2])    ) 
        self.total_features = total_features
        self.plot_centroids(**kwargs)
        self.tupdate({'feature_from': feature_from}, run_id_print=True,  **kwargs)
        return total_features, total_labels
    
    
    def train(self, **kwargs):
        ### ### calculate centroid  after each ephochs after a fresh training############ 
        ###################################################################
        start_time = time.time()
        debug = kwargs.get('oc_optimizer', False)
        oc_optimizer = kwargs.get('oc_optimizer')
        oc_lr = kwargs.get('oc_lr', 2)
        oc_epochs = kwargs.get('oc_epochs', 5)
        oc_wait = kwargs.get('oc_wait', 3)
        train_data, val_data,  test_data, bglog = self.get_or_generate_dataset(**kwargs)
        ### This was self.num_labels in V6. For 5000 ablation 2 was provided manually  with init of openset 
        ### hence it is the number of class in training data .
        ### used in : # Boundary Loss , Centroids_cal  
       
        log_obj = kwargs.get('bglog', BGLog)
        # ukc_label = kwargs.get('ukc_label', self.ukc_label)   #### ukc_label = self.self.designated_ukc_cls
        self.designated_ukc_cls = kwargs.get('designated_ukc_cls', log_obj.designated_ukc_cls)
        self.ukc_label = kwargs.get('ukc_label', self.designated_ukc_cls) 
        update_tracker = kwargs.get('update_tracker', True)
        save_ocmodel = kwargs.get('save_ocmodel', True)
        
        oc_optimizer_obj = self.get_optimizer(oc_optimizer, lr_rate=oc_lr)
        train_data, val_data,  test_data, bglog = self.get_or_generate_dataset(**kwargs)
        num_classes = kwargs.get('num_classes', train_data.element_spec[1].shape[1])
        self.num_classes = num_classes
        lossfunction = BoundaryLoss(num_labels=num_classes)
        ### ### why is it needed ? #####################
        # self.radius = tf.nn.softplus(lossfunction.theta)
        
        ######### get centroids #############
        # train_data=train_data, val_data=val_data, test_data=test_data, bglog=bglog,
        _, _ = self.extract_features_and_centroids(**kwargs)
        # self.centroids = self.centroids_cal(train_data, **kwargs) 
        # if debug: print('self.centroids calcualted: ', self.centroids)
        ####################################################    
        wait, best_radius, best_centroids = 0, None, None 
        for epoch in range(oc_epochs):
            ### ### calculate centroid after one more round of triaing
            # this is increasing the loss instead of decreasing in each epoch
            #self.pretrained_model.fit(data_train)
            #self.centroids = self.centroids_cal(data_train)  
            tr_loss, nb_tr_examples, nb_tr_steps = 0, 0, 0            
            for batch in tqdm(train_data):
                logseq_batch, label_batch = batch ## (32, 32, 64), (32, 4)
                batch_loss, self.radius = self.train_step(lossfunction, 
                                                     logseq_batch, label_batch, oc_optimizer_obj)
                tr_loss += batch_loss
                nb_tr_steps += 1                
            self.radius_changes.append(self.radius)
            loss = tr_loss / nb_tr_steps
            self.losses.append(tr_loss)
            ######don't store the feature with this evaluate instead use the feature extracted during centroid and feature extraction before for loop
            _, _, eval_score_train, _ = self.evaluate(train_data, debug=False,)
            self.f1_tr_lst.append(round(eval_score_train, 4))
            if val_data:
                _, _, eval_score_val, _ = self.evaluate(val_data, debug=False)
                self.f1_val_lst.append(round(eval_score_val, 4))
                print(f'epoch: {epoch+1}/{oc_epochs}, train_loss: {loss.numpy()}, F1_train: {eval_score_train} '
                      f'F1_val: {eval_score_val}')
            else:
                print(f'epoch: {epoch+1}/{oc_epochs}, train_loss: {loss.numpy()}, F1_train: {eval_score_train}')            
            
            if (eval_score_train > self.best_train_score) or (eval_score_val > self.best_val_score):
                wait = 0
                if eval_score_train > self.best_train_score:                
                    self.best_train_score = eval_score_train
                if val_data and eval_score_val > self.best_val_score:
                    self.best_val_score = eval_score_val
                best_radius = self.radius
                # best_centroids = self.centroids                
            else:    
                wait += 1
                if eval_score_train <= self.best_train_score:
                    print(f'train score not improving  going to wait state {wait}')
                if eval_score_val <= self.best_val_score:
                    print(f'val score not improving  going to wait state {wait}')                
                if wait >= oc_wait:                    
                    break
            self.epoch = epoch
        self.radius = best_radius
        # self.centroids = best_centroids 
        oc_tr_time = time.time() - start_time
        
        # kwargs.update({})
        self.plot_radius_chages(num_classes=self.num_classes, **kwargs)        
        print('classification report for training:')
        _, _, f1_weighted, f_measure = self.evaluate(train_data, ukc_label=self.designated_ukc_cls)
        #### expecting to store the feature of test data and test labels with open set classes 
        print('classification report for test data:')
        ##### hence store_features is true and plotting the centroid with the test features and test labels
        _, _, f1_weighted, f_measure = self.evaluate(test_data, ukc_label=self.designated_ukc_cls, store_features=True)
        centroid_plot_start = time.time()
        self.plot_centroids(use_labels=self.total_preds, **kwargs)
        centroid_plot_time = time.time() - centroid_plot_start
        total_oc_time = time.time() - start_time
        self.tracker.update({'oc_epochs': oc_epochs, 'oc_wait': oc_wait, 'oc_lr': oc_lr, 'oc_optimizer': oc_optimizer,
        
                            'oc_tr_time': oc_tr_time, 'centroid_plot_start': centroid_plot_start,
                            'total_oc_time': total_oc_time}, **kwargs)
        if update_tracker:
            self.update_tracker('mytest.xlsx', self.tracker)
        self.save_oc_model(**kwargs)
        
        return self.losses, self.radius_changes  
    
            
    def train_step(self, Lfunction, logseq_batch, label_batch, optimizer):       
        with tf.GradientTape() as tape:                
            # features_batch = self.model(logseq_batch, extract_feature=True)
            features_batch = self.get_pretrained_features(logseq_batch)
            loss, self.radius = Lfunction(features_batch, self.centroids, label_batch)        
            gradients = tape.gradient(loss, [self.radius])
            optimizer.apply_gradients(zip(gradients, [self.radius]))
        return loss, self.radius
       
        
    def get_pretrained_features(self, logseq_batch, **kwargs):
        if self.function_model is True:
            penultimate_layer = self.pretrained_model.layers[len(self.pretrained_model.layers) -2]
#             features = penultimate_layer.output
        else:
            ptmodel = self.get_ptmodel(**kwargs)
            batch_features = ptmodel(logseq_batch, extract_feature=True)
        self.batch_features = batch_features
        return batch_features
    
        
    def centroids_cal(self, data, **kwargs):
        ''' centroid rows and total_label rows should not varry with training and testing dataset. It should be always number of classes 
            of training dataset. self.num_classes should be used since this was captured while calling method train '''
        # num_classes = kwargs.get('num_classes', data.element_spec[1].shape[1]) 
        embedding_size = kwargs.get('embedding_size', 16)        
        centroids = tf.zeros((self.num_classes, embedding_size))
        total_labels = tf.zeros(self.num_classes)
        for batch in data:
            logseq_batch, label_batch = batch
            ## (32, 32, 64), (32, 4)            
            features = self.get_pretrained_features(logseq_batch, **kwargs)
            ## (32, 16) features - 32 sequence of line each haaving 64 characrers
            ## produces a feaure vector of dimension 16. 
            for i in range(len(label_batch)): # (32, 4) --> here length is 32
                label = label_batch[i] # label looks like [0 0 0 1]
                numeric_label = np.argmax(label) # index position of the label = 3 , so it is actually class =3
                ##total_labels = [0 0 0 0] each col representing a class 
                ## count the number for each class
                total_labels_lst = tf.unstack(total_labels)
                total_labels_lst[numeric_label] += 1 
                total_labels = tf.stack(total_labels_lst)
                centroids_lst = tf.unstack(centroids)
                centroids_lst[numeric_label] += features[i]
                centroids = tf.stack(centroids_lst)
                # self.labelled_features[numeric_label] = features[i]
                # each row index in the centroid array is a class
                # we add first identify the feature belonging to which class by the numeric_label
                # Then add all the features belonging to the class in the corresponding row of the centroid arr
        ### shape of centroids is (4, 16) whereas shape of total_labels is (1, 4)
        ### reshape the total_labels as 4,1 ==> [[0], [0], [0], [0]]==> 4 rows 
        ## so that we can divide the centroids array by the total_labels
        total_label_reshaped = tf.reshape(total_labels, (self.num_classes, 1))
        centroids /= total_label_reshaped
        # TODO: the centroid for a class is a scalar or vector ?
        return centroids 
    

    def openpredict(self, features, **kwargs):
        debug = kwargs.get('debug', True)
        # ukc_label = kwargs.get('ukc_label', 7)
        # print('self.centroid within openpredict ', self.centroids)
        logits = self.euclidean_metric(features, self.centroids)
        ####original line in pytorch ###probs, preds = F.softmax(logits.detach(), dim = 1).max(dim = 1)
        smax = tf.nn.softmax(logits, )
        preds = tf.math.argmax(smax, axis=1)
        probs = tf.reduce_max(smax, 1)            
        #######euc_dis = torch.norm(features - self.centroids[preds], 2, 1).view(-1)
        pred_centroids = tf.gather(self.centroids, indices=preds)
        euc_dis = tf.norm(features - pred_centroids, ord='euclidean', axis=1)
        # print('self.radius inside openpredict: ', self.radius)
        pred_radius = tf.gather(self.radius, indices=preds)
        pred_radius = tf.reshape(pred_radius, pred_radius.shape[0], )
        #####preds[euc_dis >= self.delta[preds]] = data.unseen_token_id
        unknown_filter = euc_dis >= pred_radius
        #convert to numpy since tensor is immutable
        unknown_filter_np = unknown_filter.numpy()
        preds_np = preds.numpy()
        preds_np[unknown_filter_np] = self.ukc_label  
        if debug:
            print('euc_dis:', euc_dis)
            print('pred_radius:', pred_radius)
            print(f'predictions with ukc_label={self.ukc_label}', preds_np)
        return preds_np
    
    
    def evaluate(self, data, **kwargs, ):
        # if hasattr(data, ukc_label)
        debug = kwargs.get('debug', True)
        zero_div = kwargs.get('zero_div', 1)
        ukc_label = kwargs.get('ukc_label')
        ##### as per opensetV6 ################
        if ukc_label is None:
            ukc_label = self.ukc_label
        else:
            self.ukc_label = ukc_label
        store_features = kwargs.get('store_features', False)
        
        # if ukc_label is None:
        #     ukc_label = self.ukc_label
        # else:
        #     self.ukc_label = ukc_label
        total_features, total_preds, total_labels = [], [], []
        num_samples = 0
        for batch in data:
            logseq_batch, label_batch = batch
            features_batch = self.get_pretrained_features(logseq_batch)
            # print('first feature batch within evaluate: ', features_batch[0][0])            
            preds_np = self.openpredict(features_batch, debug=False)
            total_preds.append(preds_np)
            total_features.append(features_batch)
            
            label_indexs = tf.math.argmax(label_batch, axis=1)
            label_index_np = label_indexs.numpy()
            total_labels.append(label_index_np)
            
            num_samples += 1
        y_pred = np.array(total_preds).flatten().tolist()        
        y_true = np.array(total_labels).flatten().tolist()
        if store_features:
            self.total_preds = y_pred
            total_features = np.array(total_features)
            total_features = np.reshape(total_features, ((total_features.shape[0] * total_features.shape[1]), total_features.shape[2])    ) 
            self.total_features = total_features
        cm = confusion_matrix(y_true, y_pred)
        acc = round(accuracy_score(y_true, y_pred) * 100, 2)
        f1_weighted = f1_score(y_true, y_pred, average="weighted", zero_division=zero_div)
        f1_macro = f1_score(y_true, y_pred, average="macro", zero_division=zero_div, )
        f1_micro = f1_score(y_true, y_pred, average="micro", zero_division=zero_div, )
        cls_report = m.classification_report(y_true, y_pred)
        f_measure = self.F_measure(cm, **kwargs)
        if debug:
            print(cm)
            print(acc)
            print(f'f1_weighted: {f1_weighted}, f1_macro: {f1_macro}, '
                  f'f1_micro: {f1_micro}, f_measure: {f_measure}')
            print(cls_report)
            # self.plot_centroids()
        self.tupdate({'f1_micro': f1_micro, 'f1_weighted': f1_weighted, 'oc_accu': acc,
                            'ukc_label': ukc_label}, run_id_print=False, **kwargs)
        return y_true, y_pred, f1_weighted, f_measure
    
    
    def get_ptmodel(self, **kwargs):
        self.ptmodel = kwargs.get('ptmodel', self.ptmodel)
        ptmodel_get_mode =  kwargs.get('ptmodel_get_mode', 'train')
        if self.ptmodel is None:
            if ptmodel_get_mode == 'import':
                self.ptmodel = self.import_ptmodel(**kwargs)
            else:
                self.ptmodel, hist, filepath = self.train_ptmodel(**kwargs) #ptmodel, hist, filepath
        return self.ptmodel
    
    
    def import_ptmodel(self, **kwargs):
        ''' after the training is completed this class will have the latest and best ptmodel_path
        if you want to load any specific saved model or load a saved model without training use the
        parameter ptmodel_path to provide the full or relative path e.g. data\ptmodel_2022-04-25_07_08_47.426755/ 
        or full path.'''
        self.ptmodel_path = kwargs.get('ptmodel_path', self.ptmodel_path)
        print(f'importing model: {self.ptmodel_path}')
        self.ptmodel = load_model(self.ptmodel_path)
        self.tupdate({'ptmodel_path': ptmodel_path, **kwargs})
        return self.ptmodel
    
    
    def get_ptmodel_arch(self, **kwargs):
        bglog = kwargs.get('bglog')
        train_data = kwargs.get('train_data')
        val_data = kwargs.get('val_data')
        chars_in_line = kwargs.get('chars_in_line', train_data.element_spec[0].shape[2])
        line_in_seq = kwargs.get('line_in_seq', train_data.element_spec[0].shape[1])
        char_embedding_size = kwargs.get('char_embedding_size', len(bglog.tk.word_index)) # if None self.vocabulary_size will be used by the LogLineEncoder         
        pt_optimizer = kwargs.get('pt_optimizer', 'adam')
        pt_loss = kwargs.get('pt_loss', 'categorical_crossentropy')
        pt_metrics = kwargs.get('pt_metrics', ['accuracy', tf.keras.metrics.Precision(),tf.keras.metrics.Recall()])
        tf_random_seed = kwargs.get('tf_random_seed', 1234 )
        embedding_size = kwargs.get('embedding_size', 16)
        # num_classes = kwargs.get('num_classes', train_data.element_spec[1].shape[1])
        tf.random.set_seed(self.tf_random_seed)
        if bglog is None or train_data is None or val_data is None:
            train_data, val_data,  test_data, bglog = self.get_bgdata(**kwargs)
        line_encoder = LogLineEncoder(bglog, chars_in_line=chars_in_line, char_embedding_size=char_embedding_size,)
        log_seqencer =  LogSeqEncoder(line_in_seq=line_in_seq, dense_neurons=embedding_size)
        ptmodel_arch = LogClassifier(line_encoder=line_encoder, seq_encoder=log_seqencer, num_classes=self.num_classes)
        ptmodel_arch.compile(optimizer=pt_optimizer, loss=pt_loss, metrics=pt_metrics)
        self.tupdate({'char_embedding_size': char_embedding_size, 'pt_optimizer': pt_optimizer, 'num_classes': self.num_classes,
                            'pt_loss': pt_loss}, **kwargs)
        return ptmodel_arch 
    
    
    def train_ptmodel(self, **kwargs, ):
        start_time = time.time()
        ptmodel = self.get_ptmodel_arch(**kwargs)
        ptmodel_name = kwargs.get('ptmodel_name', 'ptmodel')
        monitor_metric = kwargs.get('monitor_metric', 'accuracy')
        self.data_dir = kwargs.get('data_dir', self.data_dir)
        self.save_dir = kwargs.get('save_dir', self.data_dir)
        save_ptmodel = kwargs.get('save_ptmodel', True)
        pt_wait = kwargs.get('pt_wait', 3)
        pt_epochs = kwargs.get('pt_epochs', 5)
        pt_early_stop = kwargs.get('pt_early_stop', False)
        train_data, val_data,  test_data, bglog = self.get_or_generate_dataset( **kwargs)
        print(datetime.datetime.now())
        print('starting to create {} automatically'.format(ptmodel_name))
        curr_dt_time = datetime.datetime.now()
        model_name = self.ptmodel_name + '_' + str(curr_dt_time).replace(' ','_').replace(':','_') + '/'
        if not os.path.exists(self.save_dir):
            os.mkdir(self.save_dir)
        # model_name = os.path.join(self.save_dir, model_name) 
        filepath = os.path.join(self.save_dir, model_name) 
        # if not os.path.exists(model_name):
        #     os.mkdir(model_name)
        # filepath = model_name + 'model-{epoch:05d}-{loss:.5f}-{accuracy:.5f}.tf'         
        if val_data is not None:            
        #     filepath = model_name + 'model-{epoch:05d}-{loss:.5f}-{accuracy:.5f}-{val_loss:.5f}-{val_accuracy:.5f}.tf'
            monitor_metric = 'val_accuracy'
        checkpoint = ModelCheckpoint(filepath, monitor=monitor_metric, verbose=1, 
                                     save_best_only=True, save_weights_only=False, 
                                     save_format="tf",
                                     mode='auto', save_freq='epoch')

        LR = ReduceLROnPlateau(monitor=monitor_metric, factor=0.5, patience=2, cooldown=1, verbose=1)
        earlystop = EarlyStopping(monitor=monitor_metric, min_delta=0, patience=pt_wait, verbose=1)
        callbacks_list = [LR]
        if save_ptmodel:
            callbacks_list.append(checkpoint)
        if pt_early_stop:
            callbacks_list.append(earlystop)
        #     callbacks_list = [checkpoint, LR, earlystop]
        # else:
        #     callbacks_list = [ LR, earlystop] 
        print('staring pre trining')
        hist = ptmodel.fit(train_data, validation_data=val_data, epochs=pt_epochs,
                          callbacks=callbacks_list,  )        
        self.plot_pretrain_result(hist) 
        self.pretrained_model = ptmodel
        self.ptmodel_path = filepath
        pt_time = time.time() - start_time
        self.tupdate({'ptmodel_name': ptmodel_name, 'data_dir': self.data_dir, 'save_ptmodel': save_ptmodel, 'pt_wait': pt_wait,
                            'pt_epochs':pt_epochs,  'ptmodel_path': filepath,
                     'pt_time': pt_time}, run_id_print=True, **kwargs)
        return ptmodel, hist, filepath
    
    
    def get_or_generate_dataset(self, **kwargs):
        bg_class_obj = kwargs.get('bg_class_obj', BGLog)
        bglog = kwargs.get('bglog')
        train_data = kwargs.get('train_data')
        val_data = kwargs.get('val_data')
        test_data = kwargs.get('test_data')
        data_tuple = (bglog, train_data, val_data, test_data)         
        if all(data_tuple):
            print('got all the dataset')
        elif not all(data_tuple) and bg_class_obj is not None:
            train_data, val_data,  test_data, bglog = self.get_bgdata(**kwargs)
        else: ####not all(data_tuple) and  bg_class_obj is None:
            msg = f'you must either input all four of bglog, train_data, val_data, test_data or bg_class_obj to gnerate all four, received:  {data_tuple} '
            print(msg)
            raise OCException(message=msg)
        return train_data, val_data,  test_data, bglog      
      
    
    def get_bgdata(self, **kwargs ):
        bg_class_obj = kwargs.get('bg_class_obj')
        bglog = bg_class_obj(**kwargs)
        train_test = bglog.get_tensor_train_val_test(**kwargs)
        train_data, val_data,  test_data = train_test       
        bs = train_data.element_spec[0].shape[0]
        seql = train_data.element_spec[0].shape[1]
        chars = train_data.element_spec[0].shape[2]        
        self.tupdate({'batch_size': bs, 'padded_seq_len':seql, 'padded_char_len': chars, 
                             'logpath': bglog.logfile, 
                             'logfilename': bglog.logfilename, 
                             'pkl_file': bglog.full_pkl_path, 
                             'tk_file': bglog.tk_path, 
                             'load_from_pkl': bglog.load_from_pkl, 
                             'train_ratio': bglog.train_ratio, 
                             'ablation': bglog.ablation, 
                             'save_dir': bglog.save_dir, 
                             'designated_ukc_cls': bglog.designated_ukc_cls,
                             'clean_part_1': bglog.clean_part_1,
                             'clean_part_2': bglog.clean_part_2,
                             'clean_time_1': bglog.clean_time_1,
                             'clean_part_4': bglog.clean_part_4,
                             'clean_time_2': bglog.clean_time_2,
                             'clean_part_6': bglog.clean_part_6,
                             }, **kwargs)
        return train_data, val_data,  test_data, bglog
        
        
    def get_optimizer(self, optimizer, lr_rate=None):        
        if optimizer == 'nadam':
            optimizer = tf.keras.optimizers.Nadam(learning_rate=lr_rate) # does it take criterion_boundary.parameters() ??
            if lr_rate is None:
                optimizer = tf.keras.optimizers.Nadam()
        elif optimizer == 'sgd':
            optimizer = tf.keras.optimizers.SGD(learning_rate=lr_rate)
            if lr_rate is None:
                optimizer = tf.keras.optimizers.SGD()
        elif optimizer == 'rmsprop':
            optimizer = tf.keras.optimizers.RMSprop(learning_rate=lr_rate)
            if lr_rate is None:
                optimizer = tf.keras.optimizers.RMSprop()
        elif optimizer == 'adam':
            optimizer = tf.keras.optimizers.Adam(learning_rate=lr_rate)
            if lr_rate is None:
                optimizer = tf.keras.optimizers.Adam()
        else:
            print(f'unknown optimizer {optimizer}. assigning default as adam with lr_rate={lr_rate}')
            optimizer = tf.keras.optimizers.Adam(learning_rate=lr_rate)
            if lr_rate is None:                
                optimizer = tf.keras.optimizers.Adam()
        return optimizer
    
    
    def F_measure(self, cm, **kwargs):
        idx = 0
        rs, ps, fs = [], [], []
        n_class = cm.shape[0]
        for idx in range(n_class):
            TP = cm[idx][idx]
            r = TP / cm[idx].sum() if cm[idx].sum() != 0 else 0
            p = TP / cm[:, idx].sum() if cm[:, idx].sum() != 0 else 0
            f = 2 * r * p / (r + p) if (r + p) != 0 else 0
            rs.append(r * 100)
            ps.append(p * 100)
            fs.append(f * 100)
        f = np.mean(fs).round(4)
        f_seen = np.mean(fs[:-1]).round(4)
        f_unseen = round(fs[-1], 4)
        result = {}
        result['Known'] = f_seen
        result['Open'] = f_unseen
        result['F1-score'] = f
        self.tupdate({'f1Known': f_seen, 'F1Open':f_unseen, 'f1_weighted': f,                           
                             }, **kwargs)
        return result
    
    
    def plot_pretrain_result(self, hist, **kwargs):
        figsize = kwargs.get('figsize',  (20, 6))
        pre_scores = hist.history
        pre_scores = {k:pre_scores[k] for k in pre_scores.keys() if 'loss' not in k }
        pre_scores_df = pd.DataFrame(pre_scores)       
        pre_losses = hist.history        
        pre_losses = {k:pre_losses[k] for k in pre_losses.keys() if 'loss' in k }
        pre_losses_df = pd.DataFrame(pre_losses)   
        plt.figure(figsize=figsize)
        plt.subplot(1, 2, 1) ### 1 rows 2 column , first plot
        fig1 = sns.lineplot(data=pre_scores_df, )
        plt.legend(loc=0)
        fig1.set_ylabel("pre-training Scores")
        fig1.set_xlabel("Pre-training Epochs")
        plt.subplot(1, 2, 2) ### 1st row 2 column , 2nd plot
        fig2 = sns.lineplot(data=pre_losses_df)
        fig2.set_xlabel("Pre-training Epochs")
        fig2.set_ylabel("pre-training Loss")
        plt.show()
        
        
    def plot_radius_chages(self, **kwargs):
        # perplexity = kwargs.get('perplexity', 200)
        # early_exaggeration = kwargs.get('early_exaggeration', 12)
        # random_state = kwargs.get('random_state', 123)
        # tsne_lr = kwargs.get('tsne_lr', 80)
        radius_pic_size = kwargs.get('radius_pic_size', (20, 6))
        # num_classes = kwargs.get('num_classes')
        
        narr = np.array([elem.numpy() for elem in self.radius_changes])
        tnsr = tf.convert_to_tensor(narr)        
        tpose = tf.transpose(tnsr)
        radius = [tpose.numpy()[0][i] for i in range(self.num_classes)]
        losses = [elem.numpy() for elem in self.losses]
        f1_tr = np.array(self.f1_tr_lst) * 100
        val_tr = np.array(self.f1_val_lst) * 100        
        f1_scores = [f1_tr, val_tr]
        f1_scores = pd.DataFrame({'train':f1_tr, 'val': val_tr})        
        plt.figure(figsize=radius_pic_size)
        plt.subplot(1, 2, 1) # 1 row 2 column , first plot
        fig3a = sns.lineplot(data=radius)
        plt.legend(loc=0)
        fig3a.set_xlabel("Epochs")
        fig3a.set_ylabel("Radius")
        ax2 = plt.twinx()
        fig3b=sns.lineplot(data=f1_scores, color="purple", ax=ax2,)
        fig3b.set_ylabel("F1 score")
        ax2.legend(loc=1)
        plt.subplot(2, 2, 2) # # 1 row 2 column , 2nd plot
        fig4 = sns.lineplot(data=[losses])
        fig4.set_xlabel("Epochs")
        fig4.set_ylabel("Loss")        
        plt.show()
        
        
    def plot_centroids(self, **kwargs ):
        '''
        https://distill.pub/2016/misread-tsne/
        https://arxiv.org/abs/1712.09005
        https://www.jmlr.org/papers/volume9/vandermaaten08a/vandermaaten08a.pdf  - original paper
        https://stats.stackexchange.com/questions/263539/clustering-on-the-output-of-t-sne
        https://scikit-learn.org/stable/modules/generated/sklearn.manifold.TSNE.html
        
        From original paper:
        t-SNE has a computational and memory complexity that is quadratic in the number of datapoints.This makes it infeasible to apply the standard version of t-SNE to datasets that contain many more than,say, 10,000 points.
        
        It is thereforecommontoruntheoptimizationseveraltimesona datasettofindappropriatevaluesfortheparameters
        
        (1)it is unclearhow t-SNEperformsongeneraldimensionalityreductiontasks,(2)therelativelylocalnatureoft-SNEmakesit sensitive to thecurseoftheintrinsicdimensionalityofthedata,and(3)t-SNEis notguaranteedtoconvergetoa globaloptimumofitscostfunction.Below, wediscussthethreeweaknessesinmoredetail
        '''
        total_features = kwargs.get('total_features', self.total_features)        
        total_preds = kwargs.get('total_preds', self.total_preds)        
        total_labels = kwargs.get('total_labels', self.total_labels)
        use_labels = kwargs.get('use_labels', total_labels)        
        row = kwargs.get('row', 1)
        col = kwargs.get('col', 1) 
        fig = kwargs.get('fig', 1)  
        tsne_perplexity = kwargs.get('tsne_perplexity', 200)
        tsne_early_exaggeration = kwargs.get('tsne_early_exaggeration', 12)
        tsne_random_state = kwargs.get('tsne_random_state', 123)
        tsne_lr = kwargs.get('tsne_lr', 80)
        tsne_n_iter = kwargs.get('tsne_n_iter', 1000)
        tsne_n_iter_without_progress = kwargs.get('tsne_n_iter_without_progress', 300)
        feature_pic_size = kwargs.get('feature_pic_size', 20)
        centroid_pic_size = kwargs.get('centroid_pic_size', 200) 
        centroid_class_color = kwargs.get('centroid_class_color', False) 
        manual_color_map = kwargs.get('manual_color_map', False)
        centroid_black = kwargs.get('centroid_black', False)
        fixed_color_maps = np.array(["green","blue","yellow","pink","black","orange","purple",
                                     "red","beige","brown","gray","cyan","magenta"])
        
        features = np.array(total_features)
        centroids = self.centroids.numpy()
        labels = np.array(use_labels)
        tsne = TSNE(perplexity=tsne_perplexity, early_exaggeration=tsne_early_exaggeration, 
                    random_state=tsne_random_state, learning_rate=tsne_lr, n_iter=tsne_n_iter,
                   n_iter_without_progress=tsne_n_iter_without_progress)
        tout = tsne.fit_transform(features)
        cout = tsne.fit_transform(centroids)
        m_scaler = MinMaxScaler()
        s_scalar = StandardScaler()
        # scaled_tout = m_scaler.fit_transform(tout)
        # scaled_cout = m_scaler.fit_transform(cout)
        scaled_tout = s_scalar.fit_transform(tout)
        scaled_cout = s_scalar.fit_transform(cout)
        
        ax5 = plt.subplot(row, col, fig) # # 1 row 2 column , 2nd plot
        if manual_color_map:
            fig5 = ax5.scatter(scaled_tout[:, 0], scaled_tout[:, -1], c=fixed_color_maps[labels], s=20, cmap='tab10',)
        else:
            fig5 = ax5.scatter(scaled_tout[:, 0], scaled_tout[:, -1], c=labels, s=feature_pic_size, cmap='tab10', )
        legend1 = ax5.legend(*fig5.legend_elements(),
                            loc="upper left", title="Classes",  bbox_to_anchor=(1.05, 1))
        ax5.add_artist(legend1)
        
        cmap_1 = fig5.get_cmap().colors
        ccolor = np.array([cmap_1[i] for i in  range(len(centroids))])
        if centroid_class_color:
            ax5.scatter(scaled_cout[:, 0], scaled_cout[:, -1],  s=centroid_pic_size, c=ccolor, cmap='tab10', marker=r'd', edgecolors= 'k')
        elif manual_color_map and centroid_black is False:
            filtered_fixed_color_map = list(fixed_color_maps)
            filtered_fixed_color_map = filtered_fixed_color_map[:self.num_classes]
            last_index = len(filtered_fixed_color_map) - 1
            filtered_fixed_color_map[last_index] = 'red'
            print('the color map for the classes, here index postion are the class number:', filtered_fixed_color_map )
            ccolor = np.array([i for i in  range(len(centroids))])
            ax5.scatter(scaled_cout[:, 0], scaled_cout[:, -1],  s=200, c=fixed_color_maps[ccolor], cmap='tab10', marker=r'd', edgecolors= 'k')
        elif manual_color_map and centroid_black:            
            ax5.scatter(scaled_cout[:, 0], scaled_cout[:, -1],  s=200, c='black', cmap='tab10', marker=r'd', edgecolors= 'k')
        else:
            ax5.scatter(scaled_cout[:, 0], scaled_cout[:, -1],  s=centroid_pic_size, c='black', cmap='tab10', marker=r'd', edgecolors= 'k')
        ax5.set_xlabel("class features and their centroids")
        plt.show()
        
        
    def tupdate(self, data, run_id_print=False, **kwargs):
        train_data = kwargs.get('train_data')
        curr_dt_time = datetime.datetime.now()
        u = uuid.uuid1()
        run_id = str(curr_dt_time).replace(' ','_').replace(':','_') +'_'+ u.hex
        self.tracker.update({'id': run_id  })
        if run_id_print is True:
            print('run_id: ', run_id)
        self.tracker.update(data)
        # self.tracker_update(**kwargs)
        try:
            lst = list(tf.reshape(self.radius, (1, self.num_classes)).numpy()[0])
            lst = [str(i) for i in lst]
            radius = ','.join(lst)    
            oc_loss = self.losses[len(self.losses)-1].numpy()            
        except:
            radius = self.radius
            oc_loss = None       
        ndata = {'radius': radius, 'ocloss': oc_loss, 
                 'octrf1': self.best_train_score, 'ocvalf1': self.best_val_score, }       
        self.tracker.update(**ndata), 
        self.tracker.update(**data)
        self.tracker.update(**kwargs)
        return self.tracker
    
        
    def update_tracker(self, file_name:str, data:dict, file_path=None, **kwargs):   
        plist = ['bg_class_obj', 'train_data', 'val_data', 'test_data', 'bglog']        
        for k in plist:
            if k in data:
                data.pop(k)
        if file_path:
            file_name = os.path.join(file_path, file_name)
        if os.path.exists(file_name):
            wb = load_workbook(file_name)
        else:
            wb = Workbook()    
        wb.save(file_name)
        # wb.close(file_name)
        orig_df = pd.read_excel(file_name,)
        # print(orig_df.head())
        new_df = pd.DataFrame(data, index=[1])
        concat_df = pd.concat([orig_df, new_df], axis=0)
        # print(concat_df.head())
        concat_df.to_excel(file_name)
        return concat_df
    
    
    def euclidean_metric(self, a, b):
        a = tf.expand_dims(a, 1)
        b = tf.expand_dims(b, 0)
    #     logits = -((a - b)**2).sum(dim=2)
        logits = tf.math.reduce_sum(-tf.math.square(a - b), axis=2)
        return logits
    
    
    def save_oc_model(self, **kwargs):
        save_ocmodel = kwargs.get('save_ocmodel', True)
        if save_ocmodel:
            curr_dt_time = datetime.datetime.now()
            u = uuid.uuid1()
            model_id = str(curr_dt_time).replace(' ','_').replace(':','_') +'_'+ u.hex
            ocmodel_filename = kwargs.get('ocmodel_filename', model_id)
            ocmodel_save_path = kwargs.get('ocmodel_save_path', 'data')
            if not os.path.exists(ocmodel_save_path):
                os.mkdir(ocmodel_save_path)
            ocmodel_full_path = os.path.join(ocmodel_save_path, ocmodel_filename)
            with open(ocmodel_full_path, 'wb') as f:
                pickle.dump(self, f)
        if os.path.exists(ocmodel_full_path):
            self.tupdate({'ocmodel_full_path': ocmodel_full_path}, **kwargs)
            

class OCException(Exception):
    def __init__(self, message="Error occurred"):        
        self.message = message
        super().__init__(self.message)
        
        
        
'''


'''